import openpyxl
import openpyxl.utils.cell as cell_utils
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.workbook.workbook import Workbook

import src.functionality.constants.constants as constants


class OpenpyxlUtil:

    @staticmethod
    def load_excel_file(path, bool_data_only):
        """returns the workbook object corresponding the given path"""

        wb = openpyxl.load_workbook(path, data_only=bool_data_only)
        return wb

    @staticmethod
    def format_rtl_ws(*worksheets):
        """ right to left all the given sheets"""

        for ws in worksheets:
            ws.sheet_view.rightToLeft = True

    @staticmethod
    def format_column_width(ws, list_of_tuples):
        """ a list of tuples like (starting column, ending column, column width) for example
            ("A", "C", 10) or (1, 3 , 10) and apply given width to the given worksheet"""

        for tup in list_of_tuples:
            columns_lst = cell_utils.get_column_interval(tup[0], tup[1])
            for c in columns_lst:
                ws.column_dimensions[c].width = tup[2]

    @staticmethod
    def format_cell_horizontal_alignment(cell, position):
        """sets a cell horizontal alignment position must be one of
            {‘center’, ‘centerContinuous’, ‘general’, ‘distributed’, ‘left’, ‘right’, ‘fill’, ‘justify’}"""

        cell.alignment = Alignment(horizontal=position)

    @staticmethod
    def format_cell_color(cell, color_code):
        """set a cell background color"""

        fill_color = PatternFill(start_color=color_code, end_color=color_code, fill_type='solid')
        cell.fill = fill_color

    @staticmethod
    def get_top_row(ws):
        """returns a tuple including top row cell objects"""

        return next(ws.iter_rows(min_row=1, max_row=1))

    @staticmethod
    def format_global_top_row(ws):
        """sets the color and alignment for the first row of the given sheet"""

        top_row = OpenpyxlUtil.get_top_row(ws)
        for c in top_row:
            OpenpyxlUtil.format_cell_horizontal_alignment(c, constants.global_header_position)
            OpenpyxlUtil.format_cell_color(c, constants.global_header_color)

    @staticmethod
    def return_max_row(ws):
        """return max row of data area for given worksheet"""

        return ws.max_row

    @staticmethod
    def format_cell_comma_separated(cell):
        """given a cell format it as comma separated"""

        cell.number_format = '#,##0'

    @classmethod
    def format_column_comma_separated(cls, ws, columns_list, starting_row=1, ending_row=None):
        """format the column of given column list for the given worksheet as comma separated"""

        ending_row = cls.return_max_row(ws) if ending_row is None else ending_row

        for i in range(starting_row, ending_row + 1):
            for c in columns_list:
                cls.format_cell_comma_separated(ws[f"{c}{i}"])

    @staticmethod
    def format_cell_text(cell):
        """set a cell format to text"""

        cell.number_format = '@'

    @classmethod
    def format_column_text(cls, ws, columns_list, starting_row=1, ending_row=None):
        """set a column format to text"""

        ending_row = cls.return_max_row(ws) if ending_row is None else ending_row

        for i in range(starting_row, ending_row + 1):
            for c in columns_list:
                cls.format_cell_text(ws[f"{c}{i}"])

    @staticmethod
    def format_cell_rtl(cell):
        """set a cell format to RTL"""

        cell.alignment = Alignment(horizontal='right')

    @classmethod
    def format_column_rtl(cls, ws, columns_list, starting_row=1, ending_row=None):
        """set a column format to RTL"""

        ending_row = cls.return_max_row(ws) if ending_row is None else ending_row

        for i in range(starting_row, ending_row + 1):
            for c in columns_list:
                cls.format_cell_rtl(ws[f"{c}{i}"])

    @staticmethod
    def return_ws_by_index(wb, index):
        """returns the worksheet #index for the given workbook"""

        return wb.worksheets[index]

    @staticmethod
    def write_header(ws, header_list):
        """writing headers to given worksheet"""

        for i, h in enumerate(header_list):
            ws[f"{chr(65+i)}1"] = h

    @staticmethod
    def write_data(ws, data):
        """append the rows of given data to the give ws
        """

        for data_row in data:
            ws.append(data_row)  # only iterable arguments are acceptable

    @staticmethod
    def set_sheet_name(ws, name):
        """set the name of given worksheet"""

        ws.title = name

    @staticmethod
    def get_new_wb():
        """reruns a new empty workbook"""

        wb = Workbook()
        return wb

