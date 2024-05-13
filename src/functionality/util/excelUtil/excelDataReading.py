import openpyxl

from src.functionality.util.guiUtil import GUIUtil
from src.functionality.util.commonUtil import CommonUtil


class ExcelDataReading:

    @staticmethod
    def read(data_file_path, max_col, min_row=2, value_only=True):
        """return a list containing tuples representing rows of data in file"""

        wb = openpyxl.load_workbook(data_file_path)
        ws = wb.worksheets[0]
        data = []
        ws.iter_rows()
        for data_row in ws.iter_rows(min_row=min_row, max_col=max_col, values_only=value_only):
            data.append(data_row)
        return data

    @classmethod
    def choose_read(cls, max_col,  min_row=2, value_only=True):
        """let user select a file and return the data in the file, returns None if file selection is canceled"""

        data_file_path = GUIUtil.ask_open_file_name()

        if data_file_path == "":
            return None

        data = cls.read(data_file_path, max_col, min_row, value_only)
        return data

    @classmethod
    def read_reorder(cls, data_file_path, max_col, order_list, min_row=2, value_only=True):
        """reads the file from given path and returns reordered data"""

        data = cls.read(data_file_path, max_col, min_row, value_only)
        reordered_data = CommonUtil.reorder_column(data, order_list)
        return reordered_data

    @classmethod
    def choose_read_reorder(cls, max_col, order_list, min_row=2, value_only=True):
        """let user select a file and return the data in the file in the given order,
         returns None if file selection is canceled"""

        data = cls.choose_read(max_col, min_row, value_only)

        if data is None:
            return None

        reordered_data = CommonUtil.reorder_column(data, order_list)
        return reordered_data

